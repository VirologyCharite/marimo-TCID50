import marimo

__generated_with = "0.17.0"
app = marimo.App(width="medium", app_title="TCID50 calculation")


@app.cell(hide_code=True)
def _():
    import marimo as mo
    import pandas as pd
    import altair as alt
    import numpy as np
    from scipy import stats
    from sklearn.linear_model import LogisticRegression
    import os
    from pathlib import Path
    import tomllib
    import math
    import io
    import openpyxl
    return alt, io, math, mo, np, pd, stats


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    # TCID50 calculation
    ## Input and calculation
    This marimo notebooks calculates the TCID50 from a CSV file of the following structure:

    ID | Dilution | CPE | Replicates 
    ---|----------|-----|----------
    1  |10        | 0   | 8
    1  |100       | 1   | 8
    1  |1000      | 6   | 8
    1  |10000     | 8   | 8
    1  |100000 | 8 | 8 
    2  |10|0|8
    ||...|

    where "ID" is a a unique value to identify the sample, "Dilution" is the dilution factor, "CPE" is the number of wells showing cyptopathic effect and "Replicates" is the total number of replicates for the sample and dilution.

    The script will calculate the logit for the CPE probablity (linear transformation of logistic curve), determine the x-axis intercept (corresponding to p(CPE)=0.5)
    For each value without a valid regression (zero or all wells with CPE) it will set the value to 20% below or above the lowest or highest dilution. This has to be taken into account when visualizing the data.

    ## Visualization
    The script generates by default a logit(p(CPE))~log10(dilution) for each sample and saves it under the name of the input file suffixed with _logitplot.svg.
    Optionally you can provide a sample sheet formated like the following table:

    ID | x | color | facet_row | facet_col
    ---|----------|-----|----------|
      |  |  ...  | |

    where x is the column that contains the values for the x-axis, color the values for the coloration of the datapoints, facet_row and facet_col the values according to which sub-plots should be generated. Note that the column headers should not be "x", "color", "facet_row", "facet_col", but can be the names of the actual parameters such as "Timepoint", "Infection", "Treatment" etc. You can select in the settings which column should be used for each aspect of the plot

    ## Output
    Output files (basename = input file without extension):
    1. basename+"_TCID50.csv": Table with calculated TCID50 values and detection limits
    2. basename+"_logit_plots.svg": All regression plots
    3. basename+"_TCID50_overview.svg": Rough overview of log10(TCID50)/mL
    4. basename+"_TCID50_fancy.svg": Fancy plot is sample sheet and grouping variables were provided
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## Input""")
    return


@app.cell(hide_code=True)
def _(input_sample_sheet, input_tcid50, mo):
    mo.vstack(
        [
            mo.md(
                "Upload the TCID50 data file formatted as described above and, if you want fancy plots, also your sample sheet"
            ),
            mo.hstack(
                [
                    mo.vstack([mo.md("**TCID50 data file**"), input_tcid50]),
                    mo.vstack(
                        [mo.md("**sample sheet file**"), input_sample_sheet]
                    ),
                ]
            ),
        ]
    )
    return


@app.cell
def _(mo):
    input_tcid50 = mo.ui.file(filetypes=[".xlsx"])
    input_sample_sheet = mo.ui.file(filetypes=[".xlsx"])
    return input_sample_sheet, input_tcid50


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ## Settings
    Change the inputfile and volume variable and re-run all cells
    """
    )
    return


@app.cell
def _(color_col, facet_col_col, facet_row_col, mo, volume, x_col):
    mo.vstack(
        [
            mo.md("### Settings for TCID50 calculation"),
            volume,
            mo.md("### Settings for fancy plot"),
            mo.md(
                "Select which columns of the sample sheet definde the X-axis, the color of the datapoints, and faceting by row and col (last 3 optional)"
            ),
            x_col,
            color_col,
            facet_row_col,
            facet_col_col,
        ]
    )
    return


@app.cell
def _(df_sample_sheet, mo):
    x_col = mo.ui.dropdown(options=df_sample_sheet.columns, label="X column")
    color_col = mo.ui.dropdown(
        options=df_sample_sheet.columns, label="Color column"
    )
    facet_row_col = mo.ui.dropdown(
        options=df_sample_sheet.columns, label="Row facet column"
    )
    facet_col_col = mo.ui.dropdown(
        options=df_sample_sheet.columns, label="Col facet column"
    )
    return color_col, facet_col_col, facet_row_col, x_col


@app.cell
def _(mo):
    volume = mo.ui.number(start=0, stop=200, value=10, label="Volume / Well")
    return (volume,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""## TCID50 calculation""")
    return


@app.cell
def _(input_tcid50, io, np, pd, stats, volume):
    # read CPE values
    df = pd.read_excel(io.BytesIO(input_tcid50.contents()))
    df = df.set_index("ID")

    basename = "".join(input_tcid50.name().split(".")[:-1])

    # apply continuity correction to number of wells with CPE
    cpe_cont = (df["CPE"] + 0.5) / (df["Replicates"] + 1)

    # Calculate logit (for linear regression of logistic distribution)
    df["logit"] = np.log(cpe_cont / (1 - cpe_cont))

    # Apply volume to dilution
    df["Dilution"] = df["Dilution"] / float(volume.value)

    lower_limit = df.groupby("ID")["Dilution"].min()
    upper_limit = df.groupby("ID")["Dilution"].max()

    # log10 transform Dilution
    df["Dilution"] = np.log10(df["Dilution"])

    # Perform linear regression for each ID seperatly
    linreg = df.groupby("ID").apply(
        lambda x: pd.Series(stats.linregress(x["Dilution"], x["logit"])._asdict())[
            ["slope", "intercept", "intercept_stderr"]
        ],
        include_groups=False,
    )

    # Add lower and upper limit of detection
    linreg["Upper limit"] = linreg.index.map(upper_limit)
    linreg["Lower limit"] = linreg.index.map(lower_limit)
    linreg["TCID50/mL"] = 10 ** (-linreg["intercept"] / linreg["slope"])
    # Improvemen: Calculate TCID for all (resulting in +- inf for the ones without slope) and then just check whether each sample is outside the detection range
    linreg.loc[(linreg["slope"] == 0), "outside_detection"] = True
    linreg.loc[linreg["outside_detection"].isna(), "outside_detection"] = False
    linreg.loc[(linreg["intercept"] < 0) & (linreg["slope"] == 0), "TCID50/mL"] = (
        linreg["Lower limit"] * 0.8
    )
    linreg.loc[(linreg["intercept"] > 0) & (linreg["slope"] == 0), "TCID50/mL"] = (
        linreg["Upper limit"] * 1.2
    )
    linreg
    return basename, df, linreg


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ## Dose-Response curves for all samples
    Y-axis (p(CPE)) is [logit](https://de.wikipedia.org/wiki/Logit)-transformed (linearisation of sigmoidalen functions)
    """
    )
    return


@app.cell
def _(alt, basename, df, mo):
    _points = (
        alt.Chart(df)
        .mark_point()
        .encode(
            x=alt.X("Dilution:Q").title("Dilution [log10]"), y=alt.Y("logit:Q")
        )
    )
    _reg = _points.transform_regression(
        "Dilution", "logit", groupby=["ID"]
    ).mark_line(size=2)
    _hline = (
        alt.Chart()
        .mark_rule(
            color="red",
        )
        .encode(y=alt.datum(0))
    )
    _chart = (
        (_points + _reg + _hline)
        .properties(width=100, height=100)
        .facet(column="ID")
    )
    _chart.save(basename + "_logit_charts.svg")
    mo.ui.altair_chart(_chart)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ## TCID50 for all samples
    The following plot shows the log10(TCID50) for all samples. Points that fall outside the limit of detection are colored gray.
    """
    )
    return


@app.cell
def _(alt, linreg, np):
    _linreg = linreg.copy()
    _linreg["TCID50/mL"] = np.log10(_linreg["TCID50/mL"])
    _chart = (
        alt.Chart(_linreg.reset_index())
        .mark_point()
        .encode(
            x="ID",
            y=alt.Y("TCID50/mL").title("TCID50/mL (log10)").scale(zero=False),
            color=alt.when(alt.datum.outside_detection == True).then(
                alt.value("lightgray")
            ),
        )
        .properties(height=100, width=20 + 12 * len(_linreg))
    )
    _chart
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ## Visualization according to sample sheet
    The following plot attempts a visualiztion according to the sample sheet
    """
    )
    return


@app.cell
def _():
    return


@app.cell
def _(input_sample_sheet, io, linreg, np, pd):
    df_sample_sheet = pd.read_excel(io.BytesIO(input_sample_sheet.contents()))
    df_sample_sheet_merge = df_sample_sheet.merge(linreg, on="ID")
    df_sample_sheet_merge["TCID50/mL"] = np.log10(
        df_sample_sheet_merge["TCID50/mL"]
    )
    return df_sample_sheet, df_sample_sheet_merge


@app.cell
def _(
    alt,
    color_col,
    df_sample_sheet_merge,
    facet_col_col,
    facet_row_col,
    math,
    np,
    pd,
    x_col,
):
    limits = (
        df_sample_sheet_merge["Lower limit"].max(),
        df_sample_sheet_merge["Upper limit"].min(),
    )

    _domain = (
        math.floor(df_sample_sheet_merge["TCID50/mL"].min()),
        math.ceil(df_sample_sheet_merge["TCID50/mL"].max()),
    )

    _kwargs = {}
    if not color_col.value is None:
        _kwargs["color"] = (
            alt.Color(color_col.value)
            .sort(df_sample_sheet_merge[color_col.value].unique())
            .scale(scheme="dark2")
        )

    _point = (
        alt.Chart(df_sample_sheet_merge)
        .mark_point(clip=True)
        .encode(
            x=alt.X(x_col.value).sort(df_sample_sheet_merge[x_col.value].unique()),
            y=alt.Y("TCID50/mL").scale(domain=_domain),
            **_kwargs,
        )
    )
    _line = (
        alt.Chart(df_sample_sheet_merge)
        .mark_line(clip=True)
        .encode(
            x=alt.X(x_col.value).sort(df_sample_sheet_merge[x_col.value].unique()),
            y=alt.Y("mean(TCID50/mL)")
            .title("TCID50/mL (log10)")
            .scale(domain=_domain),
            **_kwargs,
        )
    )
    _lower_limit = (
        alt.Chart(
            pd.DataFrame({"y1": 0.5 * _domain[0], "y2": [np.log10(limits[0])]})
        )
        .mark_rect(opacity=0.6, clip=True, color="WhiteSmoke")
        .encode(y="y1:Q", y2="y2:Q")
    )
    _upper_limit = (
        alt.Chart(
            pd.DataFrame({"y1": [np.log10(limits[1])], "y2": 2 * _domain[1]})
        )
        .mark_rect(opacity=0.6, clip=True, color="WhiteSmoke")
        .encode(y="y1:Q", y2="y2:Q")
    )
    _kwargs = {}
    if not facet_col_col.value is None:
        _kwargs["column"] = alt.Column(facet_col_col.value)
    if not facet_row_col.value is None:
        _kwargs["row"] = alt.Column(facet_row_col.value)

    if len(_kwargs) > 0:
        _chart = (
            (_point + _line + _lower_limit + _upper_limit)
            .properties(width=200, height=150)
            .facet(**_kwargs)
            .configure_facet(spacing=10)
            .resolve_scale(x="independent")
        )
    else:
        _chart = (_point + _line + _lower_limit + _upper_limit).properties(
            width=200, height=150
        )
    _chart
    return


if __name__ == "__main__":
    app.run()
